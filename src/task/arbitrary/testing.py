import os

import openpyxl as xls
import pdfplumber
from openpyxl import Workbook


def create_workbook(path):
    workbook = Workbook()
    workbook.save(path)


if __name__ == "__main__":
    create_workbook("Data_from_PDF.xlsx")

wb = xls.load_workbook(filename="Data_from_PDF.xlsx", data_only=True)
ws = wb.active

i = 1

for x in os.listdir():

    if x.endswith(".pdf") or x.endswith(".PDF"):
        pdf = pdfplumber.open(x)
        text = ""

        for number, pageText in enumerate(pdf.pages):
            text += pageText.extract_text(x_tolerance=2, y_tolerance=2) + "\n"

        # replace datai=1
        ws.cell(row=1, column=i).value = x  # todo extract value
        # ws.cell(row = 2, column = i + 1).value = text

        n = 2
        for line in text.splitlines():
            ws.cell(row=n, column=i).value = line.replace("=", "")
            n += 1

        wb.save("Data_from_PDF.xlsx")

        i = i + 1

# print(text)
